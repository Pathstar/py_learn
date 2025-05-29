import pandas as pd
from openpyxl import Workbook

# 读取名称和描述文件
with open('名称.txt', 'r', encoding='utf-8') as f:
    names = [line.strip() for line in f.readlines()]
with open('描述.txt', 'r', encoding='utf-8') as f:
    descriptions = [line.strip() for line in f.readlines()]
# 确保名称和描述数量一致
if len(names) != len(descriptions):
    raise ValueError("名称和描述文件的行数不一致")
# 将名称和描述组合成元组列表
data = list(zip(names, descriptions))
# 排序：先按描述排序，相同描述再按名称排序
data.sort(key=lambda x: (x[1], x[0]))
# 创建Excel写入对象
wb = Workbook()
ws = wb.active
# 初始化列索引
col_idx = 1  # 从第1列开始
# 遍历排序后的数据
current_desc = None
for name, desc in data:
    if desc != current_desc:
        # 新描述，移动到下一组列
        if current_desc is not None:
            col_idx += 3  # 每组占3列
        current_desc = desc

        # 写入描述和名称
        ws.cell(row=1, column=col_idx, value=desc)  # 描述在第二列
        ws.cell(row=1, column=col_idx - 1, value="名称")  # 名称在第一列
        # ws.cell(row=1, column=col_idx + 1, = "")  # 第三列留空

        # 从第二行开始写数据
        row_idx = 2
    else:
        row_idx += 1

    # 写入数据
    ws.cell(row=row_idx, column=col_idx - 1, value=name)
    ws.cell(row=row_idx, column=col_idx, value=desc)
    ws.cell(row=row_idx, column=col_idx + 1, value="")
# 保存Excel文件
wb.save('排序结果.xlsx')
print("Excel文件已生成：排序结果.xlsx")